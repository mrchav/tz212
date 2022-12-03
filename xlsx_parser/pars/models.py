from datetime import datetime
from random import randrange

import pytz as pytz
from django.db import models

from django.db.models import Sum, F
from openpyxl import load_workbook


class BaseTable:

    #
    # Парсим загруженную пользователем таблицу и построчно добавляем в базу данные
    #
    def pars_table(table_file):
        wb = load_workbook(filename=table_file)

        table_name = f'{table_file.name} от {datetime.now(pytz.timezone("Europe/Moscow")).strftime("%d%m%y_%H%M%S_%f")}'
        table = Table.create(table_name)
        day_count = 1
        sheet = wb.active
        for row in sheet['A4':f'J{sheet.max_row}']:
            # случайно можем добавить новый день к дате
            if randrange(0, 2, 1) == 1:
                day_count += 1
                if day_count > 30: day_count = 30

            DataRow.objects.create(
                source_id=row[0].value,
                company=Company.get_or_create(row[1].value),
                table=table,
                date=datetime(2022, 12, day_count),  # случайная дата в декабре 22 года
                fact_qliq_data1=row[2].value,
                fact_qliq_data2=row[3].value,
                fact_qoil_data1=row[4].value,
                fact_qoil_data2=row[5].value,
                forecast_qliq_data1=row[6].value,
                forecast_qliq_data2=row[7].value,
                forecast_qoil_data1=row[8].value,
                forecast_qoil_data2=row[9].value
            )
        return table

    def create(table_name):
        table = Table.objects.create(name=table_name)
        return table


#
# считаем сумму данных по дням
#
class BaseStat:
    def qliq_stat_fact(table):
        return DataRow.objects.filter(table=table).values('date'). \
            annotate(total_qliq_fact=Sum(F('fact_qliq_data1') + F('fact_qliq_data2')))

    def qoil_stat_fact(table):
        return DataRow.objects.filter(table=table).values('date'). \
            annotate(total_qoil_fact=Sum(F('fact_qoil_data1') + F('fact_qoil_data2')))

    def qliq_stat_forecast(table):
        return DataRow.objects.filter(table=table).values('date'). \
            annotate(total_qliq_forecast=Sum(F('forecast_qliq_data1') + F('forecast_qliq_data2')))

    def qoil_stat_forecast(table):
        return DataRow.objects.filter(table=table).values('date'). \
            annotate(total_qoil_forecast=Sum(F('forecast_qoil_data1') + F('forecast_qoil_data2')))


class BaseCompany:
    # проверям, если первый раз встретили компанию, то создаем ее
    def get_or_create(company_name):
        company, _ = Company.objects.get_or_create(name=company_name)
        return company


class Company(models.Model, BaseCompany):
    name = models.CharField(max_length=200)

    def __str__(self):
        return self.name


class Table(models.Model, BaseTable):
    name = models.CharField(max_length=200)

    def __str__(self):
        return self.name


#
# строка с основными данными. каждой строке добавлем информацию к какой таблице она относится
#
class DataRow(models.Model):
    source_id = models.IntegerField(null=False)
    company = models.ForeignKey(Company, on_delete=models.PROTECT)
    table = models.ForeignKey(Table, on_delete=models.PROTECT)
    date = models.DateField()
    fact_qliq_data1 = models.IntegerField()
    fact_qliq_data2 = models.IntegerField()
    fact_qoil_data1 = models.IntegerField()
    fact_qoil_data2 = models.IntegerField()
    forecast_qliq_data1 = models.IntegerField()
    forecast_qliq_data2 = models.IntegerField()
    forecast_qoil_data1 = models.IntegerField()
    forecast_qoil_data2 = models.IntegerField()

    def __str__(self):
        return f'{self.source_id}) {self.company} {self.date}'
